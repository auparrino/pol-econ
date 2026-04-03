#!/usr/bin/env python
"""
Process provincial fiscal data from APNF (Administración Pública No Financiera).
Source: Sec. de Hacienda, serie_aif-apnf-2024_1.xlsx (2005-2024)

Full revenue breakdown per province:
- Tributarios de Origen Provincial (own taxes: IIBB, inmobiliario, etc.)
- Tributarios de Origen Nacional (coparticipation + other national transfers)
- Regalías (petroleum, gas, mining, hydroelectric)
- Otros No Tributarios
- Contribuciones Seguridad Social
- Venta bienes/servicios, Rentas propiedad, Transferencias

Dependency = Origen Nacional / (Origen Provincial + Regalías + Otros No Trib + Origen Nacional)
"""

import json
import os
import urllib.request
import openpyxl

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW_DIR = os.path.join(BASE, "data", "raw")
OUT_DIR = os.path.join(BASE, "src", "data")

APNF_URL = "https://www.argentina.gob.ar/sites/default/files/serie_aif-apnf-2024_1.xlsx"

# Sheet name → GeoJSON NAME_1
SHEET_MAP = {
    "Buenos Aires": "Buenos Aires",
    "Ciudad": "Ciudad de Buenos Aires",
    "Catamarca": "Catamarca",
    "Córdoba": "Córdoba",
    "Corrientes": "Corrientes",
    "Chaco": "Chaco",
    "Chubut": "Chubut",
    "Entre Ríos": "Entre Ríos",
    "Formosa": "Formosa",
    "Jujuy": "Jujuy",
    "La Pampa": "La Pampa",
    "La Rioja": "La Rioja",
    "Mendoza": "Mendoza",
    "Misiones": "Misiones",
    "Neuquén": "Neuquén",
    "Río Negro": "Río Negro",
    "Salta": "Salta",
    "San Juan": "San Juan",
    "San Luis": "San Luis",
    "Santa Cruz": "Santa Cruz",
    "Santa Fe": "Santa Fe",
    "Santiago del  Estero": "Santiago del Estero",  # double space in sheet name
    "Tucumán": "Tucumán",
    "Tierra del Fuego": "Tierra del Fuego",
}

# Row labels we care about (matched by substring)
ROW_KEYS = {
    "De Orígen Provincial": "own_taxes",
    "De Orígen Nacional": "national_transfers",  # includes copart + other
    "Distribución Secundaria Neta": "coparticipation",
    "Regalías": "royalties",
    "Otros No Tributarios": "other_nontax",
    "Contribuciones a la Seguridad Social": "social_security",
    "Vta.Bienes y Serv": "sales",
    "Rentas de la Propiedad": "property_income",
    "Transferencias Corrientes": "transfers_current",
    "I. INGRESOS CORRIENTES": "total_current_revenue",
    "II. GASTOS CORRIENTES": "total_current_expenditure",
    "Personal": "personnel",
}


def download_if_needed(url, dest, force=False):
    if os.path.exists(dest) and not force:
        print(f"  Using cached: {dest}")
        return
    os.makedirs(os.path.dirname(dest), exist_ok=True)
    print(f"  Downloading {url}...")
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req) as resp:
        data = resp.read()
        with open(dest, "wb") as f:
            f.write(data)
    print(f"  Saved {len(data)/1024:.0f} KB")


def safe_float(v):
    try:
        return float(v) if v is not None else 0
    except (ValueError, TypeError):
        return 0


def parse_province_sheet(sheet):
    """Parse a province sheet from the APNF workbook."""
    # Find years row (row 9 usually has 'CONCEPTO' in col A, years in cols B+)
    years = []
    year_row = None
    for row in range(1, 15):
        val = sheet.cell(row, 1).value
        if val and "CONCEPTO" in str(val).upper():
            year_row = row
            for col in range(2, sheet.max_column + 1):
                y = sheet.cell(row, col).value
                if y and isinstance(y, (int, float)):
                    years.append((col, int(y)))
            break

    if not years or not year_row:
        return None

    # Map row labels to keys
    row_map = {}
    for row in range(year_row + 1, sheet.max_row + 1):
        label = str(sheet.cell(row, 1).value or "").strip()
        for pattern, key in ROW_KEYS.items():
            if pattern in label:
                # Only map first occurrence (avoid duplicates like "Rentas de la Propiedad" in expenses)
                if key not in row_map:
                    row_map[key] = row
                # Special: "Transferencias Corrientes" appears in both income and expenses
                # We want the income one (before GASTOS)
                if key == "transfers_current" and "total_current_expenditure" in row_map:
                    pass  # skip expense side
                break

    # Extract data per year
    year_data = []
    for col, year in years:
        entry = {"year": year}
        for key, row in row_map.items():
            entry[key] = safe_float(sheet.cell(row, col).value)
        year_data.append(entry)

    return year_data


def process():
    print("Processing APNF fiscal data...")

    apnf_path = os.path.join(RAW_DIR, "apnf_2024.xlsx")
    download_if_needed(APNF_URL, apnf_path)

    wb = openpyxl.load_workbook(apnf_path, data_only=True)

    provinces = []
    for sheet_name, geo_name in SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: Sheet '{sheet_name}' not found, skipping")
            continue

        sheet = wb[sheet_name]
        year_data = parse_province_sheet(sheet)
        if not year_data:
            print(f"  WARNING: Could not parse '{sheet_name}'")
            continue

        # Latest year
        latest = year_data[-1]

        # Compute dependency
        own = latest.get("own_taxes", 0)
        royalties = latest.get("royalties", 0)
        other_nontax = latest.get("other_nontax", 0)
        national = latest.get("national_transfers", 0)
        copart = latest.get("coparticipation", 0)

        # Own revenue = provincial taxes + royalties + other non-tax
        own_total = own + royalties + other_nontax
        # Total fiscal revenue (excluding social security, sales, property which are smaller)
        fiscal_total = own_total + national
        dependency = (national / fiscal_total * 100) if fiscal_total > 0 else 0

        entry = {
            "province": geo_name,
            "year": latest["year"],
            "ownTaxes": round(own, 1),
            "royalties": round(royalties, 1),
            "otherNonTax": round(other_nontax, 1),
            "ownTotal": round(own_total, 1),
            "nationalTransfers": round(national, 1),
            "coparticipation": round(copart, 1),
            "totalRevenue": round(fiscal_total, 1),
            "dependency": round(dependency, 1),
            "totalCurrentRevenue": round(latest.get("total_current_revenue", 0), 1),
            "totalCurrentExpenditure": round(latest.get("total_current_expenditure", 0), 1),
            "personnel": round(latest.get("personnel", 0), 1),
        }

        # Time series
        ts = []
        for yd in year_data:
            o = yd.get("own_taxes", 0) + yd.get("royalties", 0) + yd.get("other_nontax", 0)
            n = yd.get("national_transfers", 0)
            tot = o + n
            dep = (n / tot * 100) if tot > 0 else 0
            ts.append({
                "year": yd["year"],
                "own": round(o, 1),
                "transfers": round(n, 1),
                "royalties": round(yd.get("royalties", 0), 1),
                "dependency": round(dep, 1),
            })
        entry["timeSeries"] = ts

        provinces.append(entry)

    result = {
        "lastUpdated": str(max(p["year"] for p in provinces)),
        "source": "APNF - Ejecución Presupuestaria Provincial (Sec. de Hacienda, 2005-2024)",
        "formula": "Dependencia = Transferencias Nacionales / (Rec. Propios + Regalías + Otros No Trib. + Transf. Nacionales)",
        "provinces": provinces,
    }

    out_path = os.path.join(OUT_DIR, "dnap_fiscal.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False)
    print(f"  Wrote {len(provinces)} provinces to {out_path}")


if __name__ == "__main__":
    process()
    print("Done.")
