"""Build src/data/dnap_empleo_provincial.json from the primary DNAP xlsx
(Ocupación y salarios — AC+OD+CE).

Source: https://www.argentina.gob.ar/economia/sechacienda/coordinacion-fiscal-provincial/ejecucion-presupuestaria-provincial/ocupacion-y-salarios
File:   https://www.argentina.gob.ar/sites/default/files/ejec-presupuestaria-ocupacion_salarios_1.xlsx

Scope (literal from xlsx 2024 sheet):
  "ADMINISTRACION CENTRAL, ORGANISMOS DESCENTRALIZADOS Y CUENTAS ESPECIALES"
  24 jurisdictions. Denominator: INDEC projection 2022-2040 (per sheet footnote).
  Excludes: municipal, national, state enterprises, pension funds.

Columns per year sheet:
  B=JURISDICCIONES, C=GASTO EN PERSONAL (miles $), D=PLANTA OCUPADA,
  E=GASTO MEDIO MENSUAL ($), F=HABITANTES, G=EMPLEADOS/1000 HAB.
"""
from __future__ import annotations
import json
import unicodedata
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "scripts" / "raw" / "dnap_ocupacion_salarios.xlsx"
OUT = ROOT / "src" / "data" / "dnap_empleo_provincial.json"

SOURCE_URL = (
    "https://www.argentina.gob.ar/economia/sechacienda/coordinacion-fiscal-provincial/"
    "ejecucion-presupuestaria-provincial/ocupacion-y-salarios"
)
FILE_URL = (
    "https://www.argentina.gob.ar/sites/default/files/"
    "ejec-presupuestaria-ocupacion_salarios_1.xlsx"
)

# Canonical display names (matches governors.js / ProvincePanel lookups)
NAME_MAP = {
    "g.c.b.a.": "Ciudad de Buenos Aires",
    "buenos aires": "Buenos Aires",
    "catamarca": "Catamarca",
    "cordoba": "Córdoba",
    "corrientes": "Corrientes",
    "chaco": "Chaco",
    "chubut": "Chubut",
    "entre rios": "Entre Ríos",
    "formosa": "Formosa",
    "jujuy": "Jujuy",
    "la pampa": "La Pampa",
    "la rioja": "La Rioja",
    "mendoza": "Mendoza",
    "misiones": "Misiones",
    "neuquen": "Neuquén",
    "rio negro": "Río Negro",
    "salta": "Salta",
    "san juan": "San Juan",
    "san luis": "San Luis",
    "santa cruz": "Santa Cruz",
    "santa fe": "Santa Fe",
    "santiago del estero": "Santiago del Estero",
    "tucuman": "Tucumán",
    "tierra del fuego": "Tierra del Fuego",
}


def norm(s: str) -> str:
    s = (s or "").strip().lower()
    # strip trailing annotations like (*), (**), (***)
    for tag in ("(*)", "(**)", "(***)"):
        s = s.replace(tag, "")
    s = s.strip()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.strip()


def parse_sheet(ws):
    """Return list of {province, employees, ratio, population, personalSpendThousands, avgMonthlySalary} plus TOTAL row."""
    rows = []
    total = None
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if len(r) < 7:
            continue
        label = r[1]
        if not isinstance(label, str):
            continue
        key = norm(label)
        if key == "total":
            total = {
                "employees": _int(r[3]),
                "ratio": _float(r[6]),
                "population": _int(r[5]),
                "personalSpendThousands": _float(r[2]),
                "avgMonthlySalary": _float(r[4]),
            }
            continue
        if key not in NAME_MAP:
            continue
        rows.append({
            "province": NAME_MAP[key],
            "employees": _int(r[3]),
            "ratio": _float(r[6]),
            "population": _int(r[5]),
            "personalSpendThousands": _float(r[2]),
            "avgMonthlySalary": _float(r[4]),
        })
    return rows, total


def _int(v):
    if v is None:
        return None
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return None


def _float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    year_sheets = [s for s in wb.sheetnames if s.isdigit()]
    year_sheets.sort(key=int)
    latest_year = int(year_sheets[-1])

    # Current snapshot from latest year
    rows, total = parse_sheet(wb[str(latest_year)])
    # Sort by ratio ascending (slimmest first, matches prior convention)
    rows.sort(key=lambda x: x["ratio"] if x["ratio"] is not None else 1e9)

    # Historical series per province (all years available)
    by_prov = {r["province"]: [] for r in rows}
    for y in year_sheets:
        year_rows, _ = parse_sheet(wb[y])
        for pr in year_rows:
            if pr["province"] in by_prov:
                by_prov[pr["province"]].append({
                    "year": int(y),
                    "employees": pr["employees"],
                    "ratio": round(pr["ratio"], 2) if pr["ratio"] is not None else None,
                })

    # National series
    national_series = []
    for y in year_sheets:
        _, tot = parse_sheet(wb[y])
        if tot:
            national_series.append({
                "year": int(y),
                "employees": tot["employees"],
                "ratio": round(tot["ratio"], 2) if tot["ratio"] is not None else None,
            })

    out = {
        "source": "DNAP — Ocupación y gastos salariales provinciales",
        "sourceUrl": SOURCE_URL,
        "fileUrl": FILE_URL,
        "year": latest_year,
        "scope": "Cargos ocupados (no personas) en Administración Central + Organismos Descentralizados + Cuentas Especiales provinciales, al mes de diciembre. Incluye los 3 poderes (ejecutivo, legislativo, judicial). Excluye Instituciones de Seguridad Social (cajas), Fondos Fiduciarios, organismos públicos con carácter empresarial, contratos/becas/pasantías imputados a otras partidas, y por construcción no cubre empleo municipal ni nacional.",
        "scopeDoc": "DAIPEF-DNCFP (2014) 'El Empleo Público en las Provincias Argentinas', nota al pie 2 y §1. https://www.argentina.gob.ar/sites/default/files/el_empleo_publico_en_las_provincias_argentinas.pdf",
        "populationSource": "INDEC — Proyecciones por jurisdicción 2022-2040",
        "national": {
            "employees": total["employees"],
            "ratioPer1000": round(total["ratio"], 2),
            "population": total["population"],
            "personalSpendingThousandsARS": round(total["personalSpendThousands"]),
            "avgMonthlySalaryARS": round(total["avgMonthlySalary"]),
        },
        "provinces": [
            {
                "province": r["province"],
                "employees": r["employees"],
                "ratioPer1000": round(r["ratio"], 2),
                "population": r["population"],
                "personalSpendingThousandsARS": round(r["personalSpendThousands"]),
                "avgMonthlySalaryARS": round(r["avgMonthlySalary"]),
            }
            for r in rows
        ],
        "timeSeries": {
            "national": national_series,
            "provinces": by_prov,
        },
    }

    OUT.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {len(rows)} provinces x {len(year_sheets)} years -> {OUT}")
    print(f"National {latest_year}: {total['employees']:,} empleados, {total['ratio']:.1f}/1000")
    print("Top 5 ratios:")
    for r in sorted(rows, key=lambda x: -x["ratio"])[:5]:
        name = r['province'].encode('ascii', 'replace').decode('ascii')
        print(f"  {name:<25} {r['ratio']:.1f}/1000  {r['employees']:,} empl.  salario medio ${r['avgMonthlySalary']:,.0f}")


if __name__ == "__main__":
    main()
